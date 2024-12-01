# Script to generate warm and fuzzies for every person in the contact list because everyone deserves to feel loved <3

# Use the Contacts list to generate generic warm and fuzzies for everyone.
# This is the same Contacts list that is used later in the process, to attach
# the warm and fuzzies to each email and send to each person.
# We do this step first, because the script `process.py` doesn't generate a 
# warm and fuzzy for a person if they don't have one in `responses.xlsx`. Hence, 
# this script generates a generic one first.

from openpyxl import load_workbook

def generate(contacts_filename="contacts.xlsx"):
    # Replace 'contacts.xlsx' with your own file
    workbook = load_workbook(filename=contacts_filename)
    sheet = workbook.active

    WARM_AND_FUZZIES = {}

    # Loop through each row in the sheet, starting from the second row to skip headers
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        name = row[1]  # Assuming the 'Name' is in the second column (index 1)
        if name:
            # Add a warm and fuzzy message for each name
            WARM_AND_FUZZIES[name.strip().upper()] = [
                (f"Hi {name.split()[0]}, thanks for being a part of CSESoc this year! "
                "You are awesome and we appreciate all your hard work and dedication. "
                "We couldn't have done it without you. "
                "We hope you had a great time, and that you achieved your goals and made some great memories. \n"
                "All the best for what the future holds."
                "From your lovely 2024 exec team.")
            ]

    return WARM_AND_FUZZIES